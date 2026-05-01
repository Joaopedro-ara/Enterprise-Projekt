from titanflow_enterprise.Mitarbeiter.db_Mitarbeiter import Datenbank
from openpyxl import Workbook
## ERP_MOdul für das Modul Des Lager wo wir Lagerbestände Einfügen können,Lagerbestände sehen können,Wert des lagers
#um abschliefßen die def  funktionen verbindung dung zum App2.py wo wir

class Lagermanger:
    def __init__(self): # verbindung zu datenbanbk herstellen
        self.db=Datenbank()
        self.cursor=self.db.cursor

    def Artikelnummer_pruefen(self,artikelnummer):
        #prüft ob die angegeben Artiekle bereist im lagesbestand ist
        sql="Select Artikelnummer from Lagerbestand where Artikelnummer=%s" #hier suche ich  den Artikelnummer raus
        self.cursor.execute(sql,(artikelnummer,))
        return self.cursor.fetchone() is not None

    def Artikel_anlegen(self,artikelnummer,bezeichnung,kategorie,menge,einheit,preis,lagerort):
        # Prüft, ob die Artikelnummer bereits existiert
        # Falls ja, wird False zurückgegeben (Artikel darf nicht doppelt angelegt werden)
        if self.Artikelnummer_pruefen(artikelnummer):
            return False
        else:
            sql=("Insert into Lagerbestand(Artikelnummer,Bezeichnung,Kategorie,Menge,Einheit,Preis,Lagerort) Values "
                 "(%s,%s,%s,%s,%s,%s,%s) ") # Einfügen der werte in der datenbank
            val=(artikelnummer,bezeichnung,kategorie,menge,einheit,preis,lagerort) # values geben zu den Tabellen
            self.cursor.execute(sql,val)
            self.db.connection.commit()
            return "Produkte wurden erfolgreich im Lager hinzugefügt"

    def bestand_abrufen_excel(self,export_option,kategorie=None,lagerort=None):
        #export_option= alle/gefiltert; kategorie-> Wenn es nach kategorie exportiert wird, lagerort nach lagerort gefiltett wird
        if export_option=="alle":
            sql="Select Artikelnummer,Bezeichnung,Kategorie,Menge,Einheit,Preis,Lagerort from Lagerbestand"
            self.cursor.execute(sql)
            daten=self.cursor.fetchall()
            if not daten: # Prüfen ob daten leer sind
                return "Keine Artikel gefunden"
            else:
                dateiname="Alle_Bestaende.xlsx"
                self.excel_erstellen(daten,dateiname)
                return f" {dateiname} wurde erstellt"
        elif export_option=="gefiltert" and kategorie:
            sql=("Select Artikelnummer,Bezeichnung,Kategorie,Menge,Einheit,Preis,Lagerort from Lagerbestand "
                 "Where Kategorie=%s")
            self.cursor.execute(sql,(kategorie,))
            daten=self.cursor.fetchall()
            if not daten:
                return "Keine Daten verfügbar"
            else:
                dateiname="Alle_Bestande_nach_Kategorien.xlsx"
                self.excel_erstellen(daten,dateiname)
                return f"{dateiname} Wurde  erfolgreich hergestellt"

        elif export_option=="gefiltert" and lagerort:
            sql = ("Select Artikelnummer,Bezeichnung,Kategorie,Menge,Einheit,Preis,Lagerort from Lagerbestand "
                   "Where Lagerort=%s")
            self.cursor.execute(sql,(lagerort,))
            daten = self.cursor.fetchall()
            if not daten:
                return "Keine Daten verfügbar"
            else:
                dateiname = "Alle_Bestande_nach_Lagerort.xlsx"
                self.excel_erstellen(daten, dateiname)
                return f"{dateiname} Wurde  erfolgreich hergestellt"
        else:
            return "Ungültige Auswahl"

    def excel_erstellen(self,daten,dateiname):
        wb = Workbook()  # Workbook= kompllte neue Exel-Datei eid erzeugt
        ws = wb.active  # Aktivieren von workbook also Aktives Blatt holen
        ws.title = "Alle_Bestaende"
        # hier definieren wir den überschriften
        headers = ["Artikelnummer", "Bezeichnung", "Kategorie", "Menge", "Einheit", "Preis (€)", "Lagerort"]
        from openpyxl.styles import Font, numbers  # für Fettschrift und zahlenfaktor
        for idx, header in enumerate(headers):
            spalte = chr(65 + idx)  # 0->A,1->B
            zelle = f"{spalte}1"  # A1,B1,A2
            ws[zelle].value = header  # Überschrift eintragen
            ws[zelle].font = Font(bold=True)  # Fett setzen

        # daten eintragen
        for row_idx, row_data in enumerate(daten, start=2):  # start=2 ,Weil erste zeile übershrift
            #  # row_data = (Artikelnummer, Bezeichnung, Kategorie, Menge, Einheit, Preis, Lagerort)
            ws[f"A{row_idx}"].value = row_data[0]  # Artikelnumer
            ws[f"B{row_idx}"].value = row_data[1]  # Bereichnung
            ws[f"C{row_idx}"].value = row_data[2]  # Kategorie
            ws[f"D{row_idx}"].value = row_data[3]  # Menge
            ws[f"D{row_idx}"].number_format = numbers.FORMAT_NUMBER  # Als Zahl formatieren
            ws[f"E{row_idx}"].value = row_data[4]  # Einheit
            ws[f"F{row_idx}"].value = row_data[5]  # Preis
            ws[f"F{row_idx}"].number_format= "€#,##0.00"  # Preis als Euro formatieren
            ws[f"G{row_idx}"].value = row_data[6]  # Lagerort

            # Datei speichern
        wb.save(dateiname)
    def bestand_abrzufen(self):
        #Alle Artiekl selecte und die Arteil Aufturefen
        sql="Select Artikelnummer,Bezeichnung,Kategorie,Menge,Einheit,Preis,Lagerort from Lagerbestand"
        self.cursor.execute(sql)
        daten=self.cursor.fetchall()
        return daten
    def bestand_andern(self,artikelnummer,neue_menge=None,diferenz=None):
        # Prüfen ob der Artikel existiert
        if not self.Artikelnummer_pruefen(artikelnummer):
            return False
        # prüfen ob mindesten ein Parameter gesetzt ist
        if neue_menge is None and diferenz is None:
            return False
        try:
            # Prüft, ob eine neue Menge übergeben wurde.
            # Falls ja, wird die Menge im Lagerbestand für die entsprechende Artikelnummer aktualisiert
            if neue_menge is not None:
                sql="Update Lagerbestand Set Menge=%s Where Artikelnummer=%s"
                werte=(neue_menge,artikelnummer)
            else:
                # Falls keine neue Menge vorhanden ist, wird die aktuelle Menge im Lagerbestand
                # um den Wert 'differenz' erhöht
                sql=" Update Lagerbestand Set Menge=Menge+%s Where Artikelnummer=%s"
                werte=(diferenz,artikelnummer)
            self.cursor.execute(sql,werte)
            self.db.connection.commit()
            return self.cursor.rowcount >0 # True,wenn eine zelle geändert habe
        except Exception as e:
            print("Fehler beim ändern des bestabds",e)
            return False

    def bestand_warnung_abrufen(self,grenze):# prarameter grenze einsetzen unm zu sehen welcher wert dan kritsiche sein so
        sql="Select Artikelnummer,Bezeichnung,Kategorie,Menge,Einheit,Preis,Lagerort  from Lagerbestand where Menge <= %s"
        self.cursor.execute(sql,(grenze,))
        daten=self.cursor.fetchall()
        if not daten:
            return "Keine kritsche zustände"
        return daten

    def lagerwert_berechnen(self):
        #lagerberecung also wert in den ERP_HTML
        sql="Select Menge, Preis from Lagerbestand"
        self.cursor.execute(sql)
        daten=self.cursor.fetchall()
        gesamtwert=0
        for menge,preis in daten:
            wert=menge*preis
            gesamtwert +=wert
        return gesamtwert

    def lagerberechnung_erp(self):
        #Lagerneberechung in der Datenbank selber
        sql=" Select Sum(Menge*Preis) from Lagerbestand"
        self.cursor.execute(sql)
        daten=self.cursor.fetchone()
        if daten[0] is None:
            return 0

        return daten

    def bestand_anzeigen_data(self):
        sql="Select Artikelnummer, Bezeichnung,Kategorie,Menge,Einheit,Preis,Lagerort From Lagerbestand"
        self.cursor.execute(sql)
        data=self.cursor.fetchall()
        if data is None:
            return "keine daten gefunden"
        return data

    def lagerwert_pro_ort(self):
        #prüfe nach Lagerort wie viel lagerwert pro Lager ort befindent
        sql="Select Lagerort,Sum(Menge*Preis) from Lagerbestand group by Lagerort"
        self.cursor.execute(sql)
        data=self.cursor.fetchall()
        if data is None:
            return "Keine Daten gefunden"
        return data












